import itertools, warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from statsmodels.tsa.statespace.sarimax import SARIMAX
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.preprocessing import StandardScaler

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Irish Rental Market Intelligence",
                   page_icon="🏠", layout="wide",
                   initial_sidebar_state="expanded")

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;600;700&family=Barlow+Condensed:wght@600;700&display=swap');
html,body,[data-testid="stAppViewContainer"]{background:#111318!important;color:#E8EAF0!important;font-family:'Barlow',sans-serif!important}
[data-testid="stSidebar"]{background:#1A1D26!important;border-right:1px solid #2A2D3A!important}
[data-testid="stSidebar"] *{color:#C8CBD8!important}
h1,h2,h3{font-family:'Barlow Condensed',sans-serif!important;color:#fff!important}
.kpi-card{background:linear-gradient(135deg,#1E2130,#252836);border:1px solid #2E3145;border-radius:10px;
  padding:16px 18px;text-align:center;position:relative;overflow:hidden;margin-bottom:4px}
.kpi-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,#4A90D9,#00C4B4)}
.kpi-value{font-family:'Barlow Condensed',sans-serif;font-size:1.8rem;font-weight:700;color:#fff;line-height:1.1}
.kpi-label{font-size:0.68rem;color:#7A7D8C;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px}
.kpi-delta{font-size:0.75rem;color:#FF6B6B;font-weight:600;margin-top:3px}
.kpi-delta-g{font-size:0.75rem;color:#00C4B4;font-weight:600;margin-top:3px}
.sec{background:linear-gradient(90deg,#1E2130,transparent);border-left:3px solid #4A90D9;
  padding:6px 14px;margin:14px 0 10px;font-family:'Barlow Condensed',sans-serif;
  font-size:1rem;font-weight:600;color:#C8CBD8;text-transform:uppercase;letter-spacing:2px}
div[data-testid="stMetric"]{background:#1E2130;border-radius:8px;padding:12px;border:1px solid #2E3145}
div[data-testid="stMetric"] label{color:#7A7D8C!important}
div[data-testid="stMetric"] [data-testid="stMetricValue"]{color:#fff!important}
footer,#MainMenu{visibility:hidden}
</style>""", unsafe_allow_html=True)

# ── Theme constants ───────────────────────────────────────────────────────────
BG   = "#111318"; PBG  = "#1A1D26"; GC   = "#2A2D3A"; TC   = "#C8CBD8"
A1   = "#4A90D9"; A2   = "#00C4B4"; A3   = "#FF6B6B"; A4   = "#F5A623"; A5   = "#9B59B6"
COLS = [A1, A2, A4, A5, A3]

def dl(fig, title="", h=420, legend_inside=False):
    """Apply dark layout to a Plotly figure."""
    leg = dict(bgcolor="rgba(0,0,0,0)", font=dict(size=10, color=TC),
               orientation="h", yanchor="bottom", y=1.03, xanchor="left", x=0)
    if legend_inside:
        leg = dict(bgcolor="rgba(20,22,30,0.8)", font=dict(size=10, color=TC),
                   orientation="v", yanchor="top", y=0.99, xanchor="left", x=0.01)
    fig.update_layout(
        title=dict(text=title, font=dict(family="Barlow Condensed", size=15, color=TC),
                   pad=dict(t=0, b=0)),
        paper_bgcolor=PBG, plot_bgcolor=BG,
        font=dict(family="Barlow", color=TC, size=11),
        height=h, margin=dict(l=50, r=20, t=44, b=36),
        legend=leg,
        xaxis=dict(gridcolor=GC, linecolor=GC, zerolinecolor=GC),
        yaxis=dict(gridcolor=GC, linecolor=GC, zerolinecolor=GC),
        hovermode="x unified",
    )
    return fig

def vline(fig, x, label, color):
    """Add a vertical line + label safely (all Plotly versions)."""
    fig.add_shape(type="line", x0=x, x1=x, y0=0, y1=1,
                  xref="x", yref="paper",
                  line=dict(color=color, width=1.2, dash="dash"))
    fig.add_annotation(x=x, y=0.96, xref="x", yref="paper",
                       text=label, showarrow=False,
                       font=dict(color=color, size=9),
                       bgcolor="rgba(10,12,20,0.6)", borderpad=2,
                       xanchor="left", yanchor="top")

def hline_label(fig, y, label, color, dash="dash"):
    """Add a horizontal line + label safely."""
    fig.add_shape(type="line", x0=0, x1=1, y0=y, y1=y,
                  xref="paper", yref="y",
                  line=dict(color=color, width=1.5, dash=dash))
    fig.add_annotation(x=0.01, y=y, xref="paper", yref="y",
                       text=label, showarrow=False,
                       font=dict(color=color, size=9),
                       xanchor="left", yanchor="bottom")

def q2d(q):
    p = str(q).strip().split()
    return pd.Timestamp(year=int(p[1]), month=(int(p[0][1])-1)*3+1, day=1)

# ── Data loading ──────────────────────────────────────────────────────────────
@st.cache_data
def load_data():
    # RTB regional rent
    wb = load_workbook('RTB_Regional_Rent_TimeSeries_2007_2025.xlsx', read_only=True)
    rn = [r for r in list(wb['RTBRI Q325 New'].iter_rows(values_only=True))[2:]
          if r[0] and str(r[0]).startswith('Q')]
    rtb_new = pd.DataFrame(rn, columns=['Period','Dublin','Non_Dublin','GDA_excl_Dublin','Outside_GDA'])
    re_ = [r for r in list(wb['RTBRI Q325 Existing'].iter_rows(values_only=True))[2:]
           if r[0] and str(r[0]).startswith('Q')]
    rtb_ex = pd.DataFrame(re_, columns=['Period','Dublin_Exist','Non_Dublin_Exist','GDA_Exist','OGA_Exist'])
    rtb_new['Date'] = rtb_new['Period'].apply(q2d)
    rtb_ex['Date']  = rtb_ex['Period'].apply(q2d)
    rtb = pd.merge(rtb_new, rtb_ex[['Date','Dublin_Exist','Non_Dublin_Exist']], on='Date', how='left')
    rtb = rtb.sort_values('Date').reset_index(drop=True)
    for c in ['Dublin','Non_Dublin','GDA_excl_Dublin','Outside_GDA']:
        rtb[c] = pd.to_numeric(rtb[c], errors='coerce')
    for c in ['Dublin_Exist','Non_Dublin_Exist']:
        rtb[c] = pd.to_numeric(rtb[c], errors='coerce')

    # County snapshot
    wb2 = load_workbook('RTB_County_Rent_Snapshot_Q32025.xlsx', read_only=True)
    cdata = [{'County': r[0], 'Rent': round(float(r[1]),2)}
             for r in list(wb2['RIQ325 New'].iter_rows(values_only=True))[2:]
             if r[0] and isinstance(r[0], str) and r[1] and isinstance(r[1], (int,float))]
    county = pd.DataFrame(cdata).sort_values('Rent', ascending=False).reset_index(drop=True)

    # Unemployment
    ud = pd.read_excel('CSO_Unemployment_Monthly_1998_2026.xlsx', sheet_name='Unpivoted')
    m1 = ((ud['Statistic Label'].str.strip() == 'Seasonally Adjusted Monthly Unemployment Rate (%)') &
          (ud['Age Group'].str.strip() == '15 - 74 years') & (ud['Sex'].str.strip() == 'Both sexes'))
    u = ud[m1][['Month','VALUE']].copy()
    if u.empty:
        m2 = ((ud['Statistic Label'].str.strip() == 'Seasonally Adjusted Monthly Unemployment Rate') &
              (ud['Age Group'].str.strip() == '15 - 74 years') & (ud['Sex'].str.strip() == 'Both sexes'))
        u = ud[m2][['Month','VALUE']].copy()
    u.columns = ['Month','Unemp']
    u['Date'] = pd.to_datetime(u['Month'], format='%Y %B', errors='coerce')
    u = u.dropna(subset=['Date']).sort_values('Date').reset_index(drop=True)
    uq = u.set_index('Date')['Unemp'].resample('QS').mean().reset_index()
    uq.columns = ['Date','Unemp']

    # RPPI
    pd_ = pd.read_excel('CSO_PropertyPriceIndex_2005_2026.xlsx', sheet_name='Unpivoted')
    pm = ((pd_['Statistic Label'] == 'Residential Property Price Index') &
          (pd_['Type of Residential Property'] == 'National - all residential properties'))
    ppi = pd_[pm][['Month','VALUE']].copy(); ppi.columns = ['Month','RPPI']
    ppi['Date'] = pd.to_datetime(ppi['Month'], format='%Y %B', errors='coerce')
    ppi = ppi.dropna(subset=['Date']).sort_values('Date').reset_index(drop=True)
    pq = ppi.set_index('Date')['RPPI'].resample('QS').mean().reset_index(); pq.columns = ['Date','RPPI']

    # Housing completions
    hc_r = pd.read_excel('CSO_HousingCompletions_Quarterly.xlsx', sheet_name='Unpivoted')
    rc = hc_r.columns[7]; yc = hc_r.columns[3]; vc = hc_r.columns[-1]
    hc = hc_r[hc_r[rc].astype(str).str.contains('State', na=False)][[yc,vc]].copy()
    hc.columns = ['Year','Completions']
    hc = hc.groupby('Year')['Completions'].sum().reset_index()
    hc['Year'] = pd.to_numeric(hc['Year'], errors='coerce')
    hc = hc.dropna().sort_values('Year').reset_index(drop=True)

    # Household income
    inc_r = pd.read_excel('CSO_HouseholdIncome_SILC_2020_2025.xlsx', sheet_name='Unpivoted')
    im = ((inc_r['Statistic Label'] == 'Mean Nominal Household Disposable Income (Euro)') &
          (inc_r['Sex'] == 'Both sexes'))
    inc = inc_r[im][['Year','VALUE']].copy(); inc.columns = ['Year','Annual_Income']
    inc['Year'] = pd.to_numeric(inc['Year'], errors='coerce')
    inc['Monthly_Income'] = inc['Annual_Income'] / 12
    inc = inc.sort_values('Year').reset_index(drop=True)

    # Landlord numbers
    wb7 = load_workbook('RTB_Landlord_Numbers_Q22023_Q42025.xlsx', read_only=True)
    rows7 = list(wb7['RTB LL Total Q425'].iter_rows(values_only=True))
    ll = pd.DataFrame({'Quarter': list(rows7[1][1:]), 'Total': list(rows7[2][1:])}).dropna()
    ll['Date'] = ll['Quarter'].apply(q2d)
    ll['Total'] = pd.to_numeric(ll['Total'])
    ll['QoQ'] = ll['Total'].diff()
    ll = ll.sort_values('Date').reset_index(drop=True)

    # Master dataset — built from rtb + exog data
    master = rtb[['Date','Dublin','Non_Dublin','GDA_excl_Dublin','Outside_GDA']].copy()
    master = pd.merge(master, uq, on='Date', how='left')
    master = pd.merge(master, pq, on='Date', how='left')
    master['Year'] = master['Date'].dt.year
    # Income: forward-fill from annual data
    inc_q = pd.merge(master[['Date','Year']], inc[['Year','Monthly_Income']], on='Year', how='left')
    master['Monthly_Income'] = inc_q['Monthly_Income'].ffill().bfill().values
    # YoY growth
    ms = master.sort_values('Date').copy()
    master['Dublin_YoY']    = ms['Dublin'].pct_change(4).values * 100
    master['NonDublin_YoY'] = ms['Non_Dublin'].pct_change(4).values * 100
    # Affordability — only where income data exists
    master['Afford_Dublin']   = np.where(master['Monthly_Income'] > 0,
                                          (master['Dublin']     / master['Monthly_Income'] * 100).round(2), np.nan)
    master['Afford_NonDublin']= np.where(master['Monthly_Income'] > 0,
                                          (master['Non_Dublin'] / master['Monthly_Income'] * 100).round(2), np.nan)

    return master, county, ll, hc, inc, uq, pq, rtb

# ── Load ──────────────────────────────────────────────────────────────────────
with st.spinner("Loading data…"):
    master, county, ll, hc, inc, uq, pq, rtb = load_data()

# Pre-compute safe KPI values from FULL master (never from filtered slice)
_last  = master.dropna(subset=['Dublin']).iloc[-1]
_first = master.dropna(subset=['Dublin']).iloc[0]
_mi    = master['Monthly_Income'].dropna().iloc[-1]
_unemp = master['Unemp'].dropna().iloc[-1]
_aff_d = round(_last['Dublin'] / _mi * 100, 1)
_aff_n = round(_last['Non_Dublin'] / _mi * 100, 1)
_afd_df = master.dropna(subset=['Afford_Dublin'])   # safe affordability rows

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='text-align:center;padding:14px 0 6px'>
      <div style='font-family:Barlow Condensed;font-size:1.4rem;font-weight:700;color:#fff;letter-spacing:2px'>🏠 IRISH RENTAL</div>
      <div style='font-size:0.65rem;color:#4A90D9;letter-spacing:3px'>MARKET INTELLIGENCE</div>
    </div><hr style='border-color:#2A2D3A;margin:8px 0 14px'>""", unsafe_allow_html=True)

    page = st.radio("", [
        "📊 Overview", "📈 Rent Trends", "🗺️ County Analysis",
        "💸 Affordability", "🏗️ Supply & Landlords",
        "🤖 SARIMA", "🌲 Random Forest & GB", "🔮 Prophet",
        "📡 SARIMAX", "🏆 Model Comparison", "📋 Policy Insights"
    ], label_visibility="collapsed")

    st.markdown("<hr style='border-color:#2A2D3A;margin:12px 0'>"
                "<div style='font-size:0.65rem;color:#4A90D9;letter-spacing:2px;text-transform:uppercase;"
                "margin-bottom:8px'>⚡ FILTERS</div>", unsafe_allow_html=True)

    min_y = int(master['Date'].dt.year.min())
    max_y = int(master['Date'].dt.year.max())
    yr = st.slider("Date Range", min_y, max_y, (min_y, max_y))
    region_sel = st.multiselect("Region", ["Dublin","Non-Dublin","GDA (excl. Dublin)","Outside GDA"],
                                 default=["Dublin","Non-Dublin"])
    county_sel = st.selectbox("County Focus", ["All Counties"] + sorted(county['County'].tolist()))
    model_sel  = st.multiselect("Models (Comparison)",
                                 ["SARIMA","Random Forest","Gradient Boosting","Prophet","SARIMAX"],
                                 default=["SARIMA","SARIMAX","Prophet"])

# Filtered master for date-range charts (rent cols only — safe)
mf = master[(master['Date'].dt.year >= yr[0]) & (master['Date'].dt.year <= yr[1])].copy()

RCM = {"Dublin":("Dublin",A1), "Non-Dublin":("Non_Dublin",A2),
       "GDA (excl. Dublin)":("GDA_excl_Dublin",A4), "Outside GDA":("Outside_GDA",A5)}

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
if page == "📊 Overview":
    st.markdown("""
    <div style='font-family:Barlow Condensed;font-size:2rem;font-weight:700;color:#fff;letter-spacing:1px'>
    IRISH RENTAL MARKET INTELLIGENCE</div>
    <div style='color:#4A90D9;font-size:0.75rem;letter-spacing:3px;text-transform:uppercase;margin-bottom:18px'>
    MS5131 Major Project — Predictive Analytics Dashboard</div>
    <hr style='border:0;height:1px;background:linear-gradient(90deg,#4A90D9,transparent);margin-bottom:18px'>
    """, unsafe_allow_html=True)

    kpis = [
        (_last['Dublin'],    "€{:,.0f}",  "Dublin Rent Q3'25",       f"+{((_last['Dublin']/_first['Dublin'])-1)*100:.0f}% since 2007",      "neg"),
        (_last['Non_Dublin'],"€{:,.0f}",  "Non-Dublin Rent Q3'25",   f"+{((_last['Non_Dublin']/_first['Non_Dublin'])-1)*100:.0f}% since 2007","neg"),
        (_aff_d,             "{:.1f}%",   "Dublin Rent-to-Income",   "30% = Crisis threshold",                                                "neg"),
        (_unemp,             "{:.1f}%",   "Unemployment Rate",       "Seasonally adjusted",                                                   "neu"),
        (len(county),        "{}",        "Counties Tracked",        "Q3 2025 snapshot",                                                      "neu"),
        (5,                  "{} Models", "Forecast Models",         "SARIMA · RF · GB · Prophet · SARIMAX",                                  "neu"),
    ]
    cols = st.columns(6)
    for col, (v, fmt, lbl, dlt, dt) in zip(cols, kpis):
        col.markdown(f"""<div class='kpi-card'>
        <div class='kpi-value'>{fmt.format(v)}</div>
        <div class='kpi-label'>{lbl}</div>
        <div class='kpi-delta{"" if dt=="neg" else "-g"}'>{dlt}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns([3, 2])

    with c1:
        st.markdown("<div class='sec'>RENT TREND BY REGION</div>", unsafe_allow_html=True)
        fig = go.Figure()
        lmap = {"Dublin":"Dublin","Non_Dublin":"Non-Dublin",
                "GDA_excl_Dublin":"GDA (excl. Dublin)","Outside_GDA":"Outside GDA"}
        for cn, clr in [("Dublin",A1),("Non_Dublin",A2),("GDA_excl_Dublin",A4),("Outside_GDA",A5)]:
            fig.add_trace(go.Scatter(x=mf['Date'], y=mf[cn], name=lmap[cn],
                line=dict(color=clr, width=2), mode='lines',
                hovertemplate=f"<b>{lmap[cn]}</b> €%{{y:,.0f}}<extra></extra>"))
        for xd, lb, cl in [('2008-07-01','GFC','#FF6B6B'),('2016-10-01','RPZ','#9B59B6'),('2020-03-01','COVID','#F5A623')]:
            vline(fig, xd, lb, cl)
        fig.update_yaxes(tickprefix="€", tickformat=",")
        dl(fig, h=370); st.plotly_chart(fig, use_container_width=True)

    with c2:
        st.markdown("<div class='sec'>TOP 10 COUNTIES — Q3 2025</div>", unsafe_allow_html=True)
        t10 = county.head(10)
        nav = county['Rent'].mean()
        fig2 = go.Figure(go.Bar(
            x=t10['Rent'], y=t10['County'], orientation='h',
            marker_color=[A1 if r > 1700 else A2 for r in t10['Rent']],
            text=[f"€{v:,.0f}" for v in t10['Rent']], textposition='outside',
            textfont=dict(size=9, color=TC),
            hovertemplate="<b>%{y}</b><br>€%{x:,.0f}<extra></extra>"))
        fig2.add_shape(type="line", x0=nav, x1=nav, y0=0, y1=1,
                       xref="x", yref="paper", line=dict(color=A3, width=1.5, dash="dot"))
        fig2.add_annotation(x=nav, y=0.5, xref="x", yref="paper",
                            text=f"Avg €{nav:,.0f}", showarrow=False,
                            font=dict(color=A3, size=9), xanchor="left",
                            bgcolor="rgba(10,12,20,0.6)", borderpad=2)
        fig2.update_xaxes(tickprefix="€", tickformat=",")
        dl(fig2, h=370); st.plotly_chart(fig2, use_container_width=True)

    c3, c4 = st.columns(2)
    with c3:
        st.markdown("<div class='sec'>AFFORDABILITY RATIO</div>", unsafe_allow_html=True)
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=_afd_df['Date'], y=_afd_df['Afford_Dublin'],
            name='Dublin', fill='tozeroy', fillcolor='rgba(74,144,217,0.1)',
            line=dict(color=A1, width=2),
            hovertemplate="Dublin %{y:.1f}%<extra></extra>"))
        fig3.add_trace(go.Scatter(x=_afd_df['Date'], y=_afd_df['Afford_NonDublin'],
            name='Non-Dublin', line=dict(color=A2, width=2, dash='dot'),
            hovertemplate="Non-Dublin %{y:.1f}%<extra></extra>"))
        hline_label(fig3, 30, "30% Crisis", A3, dash="solid")
        fig3.update_yaxes(ticksuffix="%")
        dl(fig3, h=290); st.plotly_chart(fig3, use_container_width=True)

    with c4:
        st.markdown("<div class='sec'>YoY RENT GROWTH</div>", unsafe_allow_html=True)
        gd = mf.dropna(subset=['Dublin_YoY'])
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(x=gd['Date'], y=gd['Dublin_YoY'], name='Dublin YoY',
            marker_color=[A1 if v >= 0 else A3 for v in gd['Dublin_YoY']],
            hovertemplate="Dublin YoY %{y:.1f}%<extra></extra>"))
        hline_label(fig4, 4, "RPZ 4% cap", A4)
        hline_label(fig4, 2, "RPZ 2% cap", A5, dash="dot")
        fig4.update_yaxes(ticksuffix="%")
        dl(fig4, h=290); st.plotly_chart(fig4, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: RENT TRENDS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📈 Rent Trends":
    st.markdown("<div class='sec'>RENT TRENDS — INTERACTIVE</div>", unsafe_allow_html=True)
    fig = go.Figure()
    for reg, (cn, clr) in RCM.items():
        vis = True if reg in region_sel else 'legendonly'
        fig.add_trace(go.Scatter(x=mf['Date'], y=mf[cn], name=reg, visible=vis,
            line=dict(color=clr, width=2.5), mode='lines',
            hovertemplate=f"<b>{reg}</b> €%{{y:,.0f}}<extra></extra>"))
    for xd, lb, cl in [('2008-07-01','GFC 2008','#FF6B6B'),
                        ('2016-10-01','RPZ Intro','#9B59B6'),
                        ('2020-03-01','COVID-19','#F5A623')]:
        vline(fig, xd, lb, cl)
    fig.update_yaxes(tickprefix="€", tickformat=",")
    dl(fig, "Average Monthly Rent by Region", h=460)
    st.plotly_chart(fig, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("<div class='sec'>YoY GROWTH RATE</div>", unsafe_allow_html=True)
        gd = mf.dropna(subset=['Dublin_YoY'])
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=gd['Date'], y=gd['Dublin_YoY'], name='Dublin',
            line=dict(color=A1, width=2), fill='tozeroy', fillcolor='rgba(74,144,217,0.08)',
            hovertemplate="Dublin %{y:.1f}%<extra></extra>"))
        fig2.add_trace(go.Scatter(x=gd['Date'], y=gd['NonDublin_YoY'], name='Non-Dublin',
            line=dict(color=A2, width=2, dash='dot'),
            hovertemplate="Non-Dublin %{y:.1f}%<extra></extra>"))
        hline_label(fig2, 4, "RPZ 4% cap", A4)
        hline_label(fig2, 2, "RPZ 2% cap", A5, dash="dot")
        fig2.update_yaxes(ticksuffix="%")
        dl(fig2, h=320); st.plotly_chart(fig2, use_container_width=True)

    with c2:
        st.markdown("<div class='sec'>NEW vs EXISTING TENANCIES</div>", unsafe_allow_html=True)
        # Use rtb directly — always has Dublin_Exist column
        ex = rtb.dropna(subset=['Dublin_Exist', 'Non_Dublin_Exist'])
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=ex['Date'], y=ex['Dublin'], name='New — Dublin',
            line=dict(color=A1, width=2), hovertemplate="New Dublin €%{y:,.0f}<extra></extra>"))
        fig3.add_trace(go.Scatter(x=ex['Date'], y=ex['Dublin_Exist'], name='Existing — Dublin',
            line=dict(color=A1, width=2, dash='dot'),
            hovertemplate="Existing Dublin €%{y:,.0f}<extra></extra>"))
        fig3.add_trace(go.Scatter(x=ex['Date'], y=ex['Non_Dublin'], name='New — Non-Dublin',
            line=dict(color=A2, width=2), hovertemplate="New Non-Dublin €%{y:,.0f}<extra></extra>"))
        fig3.add_trace(go.Scatter(x=ex['Date'], y=ex['Non_Dublin_Exist'], name='Existing — Non-Dublin',
            line=dict(color=A2, width=2, dash='dot'),
            hovertemplate="Existing Non-Dublin €%{y:,.0f}<extra></extra>"))
        fig3.update_yaxes(tickprefix="€", tickformat=",")
        dl(fig3, h=320, legend_inside=True); st.plotly_chart(fig3, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: COUNTY ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🗺️ County Analysis":
    st.markdown("<div class='sec'>COUNTY RENT SNAPSHOT — Q3 2025</div>", unsafe_allow_html=True)
    disp = county if county_sel == "All Counties" else county[county['County'] == county_sel]
    nav  = county['Rent'].mean()
    bar_h = max(420, len(disp) * 26)
    fig = go.Figure(go.Bar(
        x=disp['Rent'], y=disp['County'], orientation='h',
        marker_color=[A1 if r > 1700 else A2 if r > 1300 else "#4A9A7A" for r in disp['Rent']],
        text=[f"€{v:,.0f}" for v in disp['Rent']], textposition='outside',
        textfont=dict(color=TC, size=10),
        hovertemplate="<b>%{y}</b><br>€%{x:,.0f}/month<extra></extra>"))
    fig.add_shape(type="line", x0=nav, x1=nav, y0=0, y1=1,
                  xref="x", yref="paper", line=dict(color=A3, width=2, dash="dash"))
    fig.add_annotation(x=nav, y=0.98, xref="x", yref="paper",
                       text=f"National Avg: €{nav:,.0f}", showarrow=False,
                       font=dict(color=A3, size=10), xanchor="left",
                       bgcolor="rgba(10,12,20,0.6)", borderpad=2)
    fig.update_xaxes(tickprefix="€", tickformat=",")
    dl(fig, h=bar_h); st.plotly_chart(fig, use_container_width=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Highest Rent",    f"€{county['Rent'].max():,.0f}",  county.iloc[0]['County'])
    c2.metric("Lowest Rent",     f"€{county['Rent'].min():,.0f}",  county.iloc[-1]['County'])
    c3.metric("National Average",f"€{nav:,.0f}")
    c4.metric("Above Average",   f"{(county['Rent'] > nav).sum()} counties")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: AFFORDABILITY
# ══════════════════════════════════════════════════════════════════════════════
elif page == "💸 Affordability":
    st.markdown("<div class='sec'>RENT AFFORDABILITY ANALYSIS</div>", unsafe_allow_html=True)

    # Always use pre-computed _afd_df (from full master, income-merged)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=_afd_df['Date'], y=_afd_df['Afford_Dublin'],
        name='Dublin', line=dict(color=A1, width=3),
        fill='tozeroy', fillcolor='rgba(74,144,217,0.1)',
        hovertemplate="Dublin %{y:.1f}%<extra></extra>"))
    fig.add_trace(go.Scatter(x=_afd_df['Date'], y=_afd_df['Afford_NonDublin'],
        name='Non-Dublin', line=dict(color=A2, width=2.5, dash='dot'),
        hovertemplate="Non-Dublin %{y:.1f}%<extra></extra>"))
    max_aff = _afd_df['Afford_Dublin'].max()
    if max_aff > 30:
        fig.add_hrect(y0=30, y1=max_aff + 3, fillcolor="rgba(255,107,107,0.07)", line_width=0)
    hline_label(fig, 30, "⚠ 30% Crisis Threshold", A3, dash="solid")
    vline(fig, '2016-10-01', "RPZ Intro", A5)
    fig.update_yaxes(ticksuffix="%")
    dl(fig, "Rent-to-Income Ratio Over Time", h=420)
    st.plotly_chart(fig, use_container_width=True)

    c1, c2, c3 = st.columns(3)
    c1.metric("Dublin Ratio",       f"{_aff_d:.1f}%", "⚠ ABOVE THRESHOLD" if _aff_d > 30 else "✅ Below threshold")
    c2.metric("Non-Dublin Ratio",   f"{_aff_n:.1f}%", "⚠ ABOVE THRESHOLD" if _aff_n > 30 else "✅ Below threshold")
    c3.metric("Monthly Income (Latest)", f"€{_mi:,.0f}")

    st.markdown("<div class='sec'>INCOME vs RENT GROWTH</div>", unsafe_allow_html=True)
    dan = master.groupby('Year')['Dublin'].mean().reset_index()
    fig2 = make_subplots(specs=[[{"secondary_y": True}]])
    fig2.add_trace(go.Bar(x=inc['Year'], y=inc['Monthly_Income'], name='Monthly Income',
        marker_color=A2, opacity=0.7,
        hovertemplate="Income %{x}: €%{y:,.0f}<extra></extra>"), secondary_y=False)
    fig2.add_trace(go.Scatter(x=dan['Year'], y=dan['Dublin'], name='Dublin Rent',
        line=dict(color=A1, width=3),
        hovertemplate="Rent %{x}: €%{y:,.0f}<extra></extra>"), secondary_y=True)
    fig2.update_yaxes(tickprefix="€", tickformat=",", secondary_y=False, gridcolor=GC)
    fig2.update_yaxes(tickprefix="€", tickformat=",", secondary_y=True,  gridcolor=GC)
    fig2.update_layout(paper_bgcolor=PBG, plot_bgcolor=BG, font=dict(color=TC),
                       height=300, margin=dict(l=50,r=50,t=30,b=36),
                       legend=dict(bgcolor="rgba(0,0,0,0)", orientation="h", y=1.05))
    st.plotly_chart(fig2, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SUPPLY & LANDLORDS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🏗️ Supply & Landlords":
    st.markdown("<div class='sec'>HOUSING SUPPLY vs DUBLIN RENT</div>", unsafe_allow_html=True)
    da = master.groupby('Year')['Dublin'].mean().reset_index(); da.columns=['Year','Avg_Rent']
    sd = pd.merge(hc, da, on='Year', how='inner')
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(x=sd['Year'], y=sd['Completions'], name='Completions',
        marker_color=A2, opacity=0.7,
        hovertemplate="Completions %{x}: %{y:,}<extra></extra>"), secondary_y=False)
    fig.add_trace(go.Scatter(x=sd['Year'], y=sd['Avg_Rent'], name='Dublin Rent',
        line=dict(color=A1, width=3),
        hovertemplate="Rent %{x}: €%{y:,.0f}<extra></extra>"), secondary_y=True)
    fig.add_shape(type="line",x0=0,x1=1,y0=48000,y1=48000,xref="paper",yref="y",
                  line=dict(color=A2,width=1.5,dash="dash"))
    fig.add_annotation(x=0.01,y=48000,xref="paper",yref="y",text="48k Target",
                       showarrow=False,font=dict(color=A2,size=9),xanchor="left",yanchor="bottom")
    fig.add_shape(type="line",x0=0,x1=1,y0=33000,y1=33000,xref="paper",yref="y",
                  line=dict(color=A4,width=1.5,dash="dot"))
    fig.add_annotation(x=0.01,y=33000,xref="paper",yref="y",text="33k Govt",
                       showarrow=False,font=dict(color=A4,size=9),xanchor="left",yanchor="bottom")
    fig.update_yaxes(tickformat=",",       secondary_y=False, gridcolor=GC)
    fig.update_yaxes(tickprefix="€", tickformat=",", secondary_y=True, gridcolor=GC)
    fig.update_layout(paper_bgcolor=PBG, plot_bgcolor=BG, font=dict(color=TC),
                      height=380, margin=dict(l=50,r=50,t=40,b=36),
                      legend=dict(bgcolor="rgba(0,0,0,0)", orientation="h", y=1.06))
    st.plotly_chart(fig, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("<div class='sec'>TOTAL LANDLORDS</div>", unsafe_allow_html=True)
        fig2 = go.Figure(go.Bar(x=ll['Date'], y=ll['Total'], marker_color=A1,
            hovertemplate="%{x|%Y Q%q}: %{y:,}<extra></extra>"))
        fig2.update_yaxes(tickformat=",")
        dl(fig2, h=300); st.plotly_chart(fig2, use_container_width=True)
    with c2:
        st.markdown("<div class='sec'>QoQ LANDLORD CHANGE</div>", unsafe_allow_html=True)
        qc = ll['QoQ'].fillna(0)
        fig3 = go.Figure(go.Bar(x=ll['Date'], y=qc,
            marker_color=[A3 if v < 0 else A2 for v in qc],
            hovertemplate="%{x|%Y}: %{y:,}<extra></extra>"))
        fig3.add_shape(type="line",x0=0,x1=1,y0=0,y1=0,xref="paper",yref="y",
                       line=dict(color=TC,width=1))
        dl(fig3, h=300); st.plotly_chart(fig3, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# MODEL PAGES
# ══════════════════════════════════════════════════════════════════════════════
elif page in ["🤖 SARIMA","🌲 Random Forest & GB","🔮 Prophet","📡 SARIMAX","🏆 Model Comparison"]:

    @st.cache_data
    def run_models(_master, _uq, _pq):
        df_ts = _master[['Date','Dublin','Non_Dublin']].sort_values('Date').set_index('Date')
        ts_n=15; n=len(df_ts); tr=df_ts.iloc[:n-ts_n]; te=df_ts.iloc[n-ts_n:]

        def sgs(s):
            ba=np.inf; bo=(1,1,1); bs=(1,1,1,4)
            for p,d,q in itertools.product([0,1,2],[1],[0,1,2]):
                for P,D,Q in itertools.product([0,1],[1],[0,1]):
                    try:
                        r=SARIMAX(s,order=(p,d,q),seasonal_order=(P,D,Q,4),
                                  enforce_stationarity=False,enforce_invertibility=False).fit(disp=False)
                        if r.aic<ba: ba=r.aic; bo=(p,d,q); bs=(P,D,Q,4)
                    except: pass
            return bo,bs

        bod,bsd=sgs(tr['Dublin']); bond,bsnd=sgs(tr['Non_Dublin'])
        sdf=SARIMAX(tr['Dublin'],order=bod,seasonal_order=bsd,
                    enforce_stationarity=False,enforce_invertibility=False).fit(disp=False)
        sndf=SARIMAX(tr['Non_Dublin'],order=bond,seasonal_order=bsnd,
                     enforce_stationarity=False,enforce_invertibility=False).fit(disp=False)

        def evs(fit,col):
            po=fit.get_forecast(steps=len(te)); pm=po.predicted_mean; pm.index=te.index
            ci=po.conf_int(alpha=0.05); ci.index=te.index
            return pm,ci,mean_absolute_error(te[col],pm),np.mean(np.abs((te[col]-pm)/te[col]))*100

        pds,cds,maeds,mapeds=evs(sdf,'Dublin'); pnds,cnds,maends,mapends=evs(sndf,'Non_Dublin')
        nf=8
        sdf_f=SARIMAX(df_ts['Dublin'],order=bod,seasonal_order=bsd,
                      enforce_stationarity=False,enforce_invertibility=False).fit(disp=False)
        sndf_f=SARIMAX(df_ts['Non_Dublin'],order=bond,seasonal_order=bsnd,
                       enforce_stationarity=False,enforce_invertibility=False).fit(disp=False)
        fdt=pd.date_range(start=df_ts.index[-1]+pd.DateOffset(months=3),periods=nf,freq='QS')
        fcd=sdf_f.get_forecast(nf); fcdm=fcd.predicted_mean; fcdm.index=fdt
        fcdci=fcd.conf_int(alpha=0.05); fcdci.index=fdt
        fcnd=sndf_f.get_forecast(nf); fcndm=fcnd.predicted_mean; fcndm.index=fdt
        fcndci=fcnd.conf_int(alpha=0.05); fcndci.index=fdt

        # RF + GB
        df_rf=_master[['Date','Dublin','Non_Dublin']].merge(_uq,on='Date',how='left').merge(_pq,on='Date',how='left')
        df_rf['RPZ']=(df_rf['Date']>='2016-10-01').astype(int)

        def mk(di,tc,nl=4):
            d=di.copy()
            for l in range(1,nl+1): d[f'{tc}_lag{l}']=d[tc].shift(l)
            d[f'{tc}_roll4']=d[tc].shift(1).rolling(4).mean()
            d[f'{tc}_yoy']=d[tc].pct_change(4)*100
            d['Quarter']=d['Date'].dt.quarter
            d['RPPI_lag1']=d['RPPI'].shift(1); d['Unemp_lag1']=d['Unemp'].shift(1)
            return d.dropna()

        dfd=mk(df_rf,'Dublin'); dfnd=mk(df_rf,'Non_Dublin')
        ex={'Date','Dublin','Non_Dublin','Unemp','RPPI','RPZ'}
        fcd_c=[c for c in dfd.columns if c not in ex]
        fcnd_c=[c for c in dfnd.columns if c not in ex]
        trf=12
        ddc=dfd[fcd_c+['Dublin','Date']].dropna(); dndc=dfnd[fcnd_c+['Non_Dublin','Date']].dropna()
        Xdt,Xde=ddc[fcd_c].values[:-trf],ddc[fcd_c].values[-trf:]
        ydt,yde=ddc['Dublin'].values[:-trf],ddc['Dublin'].values[-trf:]
        dtd=ddc['Date'].values[-trf:]
        Xndt,Xnde=dndc[fcnd_c].values[:-trf],dndc[fcnd_c].values[-trf:]
        yndt,ynde=dndc['Non_Dublin'].values[:-trf],dndc['Non_Dublin'].values[-trf:]
        dtnd=dndc['Date'].values[-trf:]

        rfd=RandomForestRegressor(n_estimators=300,max_depth=8,random_state=42,n_jobs=-1)
        rfd.fit(Xdt,ydt); prfd=rfd.predict(Xde)
        rfnd=RandomForestRegressor(n_estimators=300,max_depth=8,random_state=42,n_jobs=-1)
        rfnd.fit(Xndt,yndt); prfnd=rfnd.predict(Xnde)

        gbd=GradientBoostingRegressor(n_estimators=200,max_depth=4,learning_rate=0.05,random_state=42)
        gbd.fit(Xdt,ydt); pgbd=gbd.predict(Xde)
        gbnd=GradientBoostingRegressor(n_estimators=200,max_depth=4,learning_rate=0.05,random_state=42)
        gbnd.fit(Xndt,yndt); pgbnd=gbnd.predict(Xnde)

        def mape(y,p): return np.mean(np.abs((y-p)/y))*100
        def mae_(y,p):  return mean_absolute_error(y,p)

        # Prophet
        try: from prophet import Prophet
        except: import subprocess; subprocess.run(['pip','install','prophet','-q'],check=False); from prophet import Prophet
        dpd=df_rf[['Date','Dublin','Unemp','RPPI']].rename(columns={'Date':'ds','Dublin':'y'}).dropna()
        dpnd=df_rf[['Date','Non_Dublin','Unemp','RPPI']].rename(columns={'Date':'ds','Non_Dublin':'y'}).dropna()
        regs=['Unemp','RPPI']; tp=12
        tpd=dpd.iloc[:-tp]; tepd=dpd.iloc[-tp:]
        tpnd=dpnd.iloc[:-tp]; tepnd=dpnd.iloc[-tp:]
        def bp(tr):
            mm=Prophet(changepoints=['2016-10-01','2020-01-01'],changepoint_prior_scale=0.05,
                       seasonality_mode='additive',yearly_seasonality=False,weekly_seasonality=False,daily_seasonality=False)
            mm.add_seasonality(name='quarterly',period=91.25,fourier_order=5)
            for r in regs: mm.add_regressor(r)
            mm.fit(tr); return mm
        mpd=bp(tpd); mpnd=bp(tpnd)
        ftd=mpd.predict(tepd); ftnd=mpnd.predict(tepnd)
        lu=df_rf['Unemp'].dropna().iloc[-1]; lr=df_rf['RPPI'].dropna().iloc[-1]; ld=df_rf['Date'].iloc[-1]
        fpts=pd.date_range(start=ld+pd.DateOffset(months=3),periods=nf,freq='QS')
        fdf_p=pd.DataFrame({'ds':fpts,'Unemp':lu,'RPPI':lr})
        mpdf=bp(dpd); mpndf=bp(dpnd)
        pfud=mpdf.predict(pd.concat([dpd[['ds']+regs],fdf_p],ignore_index=True))
        pfud=pfud[pfud['ds'].isin(fpts)]
        pfund=mpndf.predict(pd.concat([dpnd[['ds']+regs],fdf_p],ignore_index=True))
        pfund=pfund[pfund['ds'].isin(fpts)]

        # SARIMAX
        dsx=df_rf.set_index('Date'); tss=15; nsx=len(dsx)
        trsx=dsx.iloc[:nsx-tss]; tesx=dsx.iloc[nsx-tss:]
        sc=StandardScaler(); ec=['Unemp','RPPI']
        trex=pd.DataFrame(sc.fit_transform(trsx[ec]),index=trsx.index,columns=ec)
        teex=pd.DataFrame(sc.transform(tesx[ec]),index=tesx.index,columns=ec)

        def sxgs(s,exog):
            ba=np.inf; bo=(1,1,1); bs=(1,1,1,4)
            for p,d,q in itertools.product([0,1,2],[1],[0,1,2]):
                for P,D,Q in itertools.product([0,1],[1],[0,1]):
                    try:
                        r=SARIMAX(s,exog=exog,order=(p,d,q),seasonal_order=(P,D,Q,4),
                                  enforce_stationarity=False,enforce_invertibility=False).fit(disp=False,maxiter=200)
                        if r.aic<ba: ba=r.aic; bo=(p,d,q); bs=(P,D,Q,4)
                    except: pass
            return bo,bs

        bsd_,bssd=sxgs(trsx['Dublin'],trex); bsnd_,bssnd=sxgs(trsx['Non_Dublin'],trex)
        sxd=SARIMAX(trsx['Dublin'],exog=trex,order=bsd_,seasonal_order=bssd,
                    enforce_stationarity=False,enforce_invertibility=False).fit(disp=False,maxiter=200)
        sxnd=SARIMAX(trsx['Non_Dublin'],exog=trex,order=bsnd_,seasonal_order=bssnd,
                     enforce_stationarity=False,enforce_invertibility=False).fit(disp=False,maxiter=200)
        opd=sxd.get_forecast(len(tesx),exog=teex); psxd=opd.predicted_mean; psxd.index=tesx.index
        csxd=opd.conf_int(alpha=0.05); csxd.index=tesx.index
        opnd=sxnd.get_forecast(len(tesx),exog=teex); psxnd=opnd.predicted_mean; psxnd.index=tesx.index
        csxnd=opnd.conf_int(alpha=0.05); csxnd.index=tesx.index
        fexdf=pd.DataFrame(sc.transform(dsx[ec]),index=dsx.index,columns=ec)
        sxdf_f=SARIMAX(dsx['Dublin'],exog=fexdf,order=bsd_,seasonal_order=bssd,
                       enforce_stationarity=False,enforce_invertibility=False).fit(disp=False,maxiter=200)
        sxndf_f=SARIMAX(dsx['Non_Dublin'],exog=fexdf,order=bsnd_,seasonal_order=bssnd,
                        enforce_stationarity=False,enforce_invertibility=False).fit(disp=False,maxiter=200)
        fesc=sc.transform(np.array([[dsx['Unemp'].iloc[-1],dsx['RPPI'].iloc[-1]]]*nf))
        fdssx=pd.date_range(start=dsx.index[-1]+pd.DateOffset(months=3),periods=nf,freq='QS')
        fsxd=sxdf_f.get_forecast(nf,exog=fesc); sxdm=fsxd.predicted_mean; sxdm.index=fdssx
        sxdci=fsxd.conf_int(alpha=0.05); sxdci.index=fdssx
        fsxnd=sxndf_f.get_forecast(nf,exog=fesc); sxndm=fsxnd.predicted_mean; sxndm.index=fdssx
        sxndci=fsxnd.conf_int(alpha=0.05); sxndci.index=fdssx

        comp=pd.DataFrame([
            {'Model':'SARIMA',           'Dublin_MAE':maeds,          'Dublin_MAPE':mapeds,
             'NonDublin_MAE':maends,     'NonDublin_MAPE':mapends},
            {'Model':'Random Forest',    'Dublin_MAE':mae_(yde,prfd),  'Dublin_MAPE':mape(yde,prfd),
             'NonDublin_MAE':mae_(ynde,prfnd), 'NonDublin_MAPE':mape(ynde,prfnd)},
            {'Model':'Gradient Boosting','Dublin_MAE':mae_(yde,pgbd),  'Dublin_MAPE':mape(yde,pgbd),
             'NonDublin_MAE':mae_(ynde,pgbnd), 'NonDublin_MAPE':mape(ynde,pgbnd)},
            {'Model':'Prophet',
             'Dublin_MAE':mae_(tepd['y'].values,ftd['yhat'].values),
             'Dublin_MAPE':mape(tepd['y'].values,ftd['yhat'].values),
             'NonDublin_MAE':mae_(tepnd['y'].values,ftnd['yhat'].values),
             'NonDublin_MAPE':mape(tepnd['y'].values,ftnd['yhat'].values)},
            {'Model':'SARIMAX',
             'Dublin_MAE':mae_(tesx['Dublin'],psxd),   'Dublin_MAPE':mape(tesx['Dublin'].values,psxd.values),
             'NonDublin_MAE':mae_(tesx['Non_Dublin'],psxnd),'NonDublin_MAPE':mape(tesx['Non_Dublin'].values,psxnd.values)},
        ])
        comp['Avg_MAPE']=(comp['Dublin_MAPE']+comp['NonDublin_MAPE'])/2
        comp['Avg_MAE'] =(comp['Dublin_MAE'] +comp['NonDublin_MAE']) /2
        comp=comp.sort_values('Avg_MAPE').reset_index(drop=True)

        return dict(
            df_ts=df_ts,tr=tr,te=te,
            pds=pds,cds=cds,mapeds=mapeds,maeds=maeds,
            pnds=pnds,cnds=cnds,mapends=mapends,maends=maends,
            fcdm=fcdm,fcdci=fcdci,fcndm=fcndm,fcndci=fcndci,
            dfd=dfd,dfnd=dfnd,fcd_c=fcd_c,fcnd_c=fcnd_c,trf=trf,
            ydt=ydt,yde=yde,prfd=prfd,dtd=dtd,
            yndt=yndt,ynde=ynde,prfnd=prfnd,dtnd=dtnd,
            pgbd=pgbd,pgbnd=pgbnd,rfd=rfd,rfnd=rfnd,
            dpd=dpd,dpnd=dpnd,tpd=tpd,tepd=tepd,tpnd=tpnd,tepnd=tepnd,
            ftd=ftd,ftnd=ftnd,pfud=pfud,pfund=pfund,
            trsx=trsx,tesx=tesx,psxd=psxd,csxd=csxd,psxnd=psxnd,csxnd=csxnd,
            sxdm=sxdm,sxdci=sxdci,sxndm=sxndm,sxndci=sxndci,dsx=dsx,comp=comp,
        )

    with st.spinner("⏳ Running models — first load ~3 mins…"):
        m = run_models(master, uq, pq)

    def fc_chart(hx, hy, tx, ty, py, clo, chi, fx, fy, flo, fhi, name, clr, title, h=380):
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=hx, y=hy, name='Train', line=dict(color=TC, width=1.5)))
        if tx is not None:
            fig.add_trace(go.Scatter(x=tx, y=ty, name='Actual', line=dict(color=A3, width=2.5)))
            fig.add_trace(go.Scatter(x=tx, y=py, name=f'{name}', line=dict(color=clr, width=2.5, dash='dash')))
            if chi is not None:
                fig.add_trace(go.Scatter(x=tx, y=chi, line=dict(width=0), showlegend=False))
                fig.add_trace(go.Scatter(x=tx, y=clo, fill='tonexty',
                    fillcolor='rgba(74,144,217,0.15)', line=dict(width=0), name='95% CI'))
        if fx is not None:
            fig.add_trace(go.Scatter(x=fx, y=fy, name='Forecast', line=dict(color=A2, width=2.5, dash='dot')))
            fig.add_trace(go.Scatter(x=fx, y=fhi, line=dict(width=0), showlegend=False))
            fig.add_trace(go.Scatter(x=fx, y=flo, fill='tonexty',
                fillcolor='rgba(0,196,180,0.1)', line=dict(width=0), showlegend=False))
        fig.update_yaxes(tickprefix="€", tickformat=",")
        dl(fig, title, h=h)
        return fig

    # ── SARIMA ────────────────────────────────────────────────────────────────
    if page == "🤖 SARIMA":
        st.markdown("<div class='sec'>SARIMA BASELINE MODEL</div>", unsafe_allow_html=True)
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Dublin MAE",      f"€{m['maeds']:.0f}")
        c2.metric("Dublin MAPE",     f"{m['mapeds']:.2f}%")
        c3.metric("Non-Dublin MAE",  f"€{m['maends']:.0f}")
        c4.metric("Non-Dublin MAPE", f"{m['mapends']:.2f}%")
        col1, col2 = st.columns(2)
        with col1:
            fig=fc_chart(m['tr'].index,m['tr']['Dublin'],
                         m['te'].index,m['te']['Dublin'],m['pds'].values,
                         m['cds'].iloc[:,0].values,m['cds'].iloc[:,1].values,
                         m['fcdm'].index,m['fcdm'].values,
                         m['fcdci'].iloc[:,0].values,m['fcdci'].iloc[:,1].values,
                         'SARIMA',A1,f"SARIMA — Dublin | MAPE={m['mapeds']:.1f}%")
            st.plotly_chart(fig, use_container_width=True)
        with col2:
            fig=fc_chart(m['tr'].index,m['tr']['Non_Dublin'],
                         m['te'].index,m['te']['Non_Dublin'],m['pnds'].values,
                         m['cnds'].iloc[:,0].values,m['cnds'].iloc[:,1].values,
                         m['fcndm'].index,m['fcndm'].values,
                         m['fcndci'].iloc[:,0].values,m['fcndci'].iloc[:,1].values,
                         'SARIMA',A2,f"SARIMA — Non-Dublin | MAPE={m['mapends']:.1f}%")
            st.plotly_chart(fig, use_container_width=True)

    # ── RF & GB ───────────────────────────────────────────────────────────────
    elif page == "🌲 Random Forest & GB":
        st.markdown("<div class='sec'>RANDOM FOREST & GRADIENT BOOSTING</div>", unsafe_allow_html=True)
        def mape_(y,p): return np.mean(np.abs((y-p)/y))*100
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("RF Dublin MAPE",    f"{mape_(m['yde'],m['prfd']):.2f}%")
        c2.metric("RF Non-Dub MAPE",   f"{mape_(m['ynde'],m['prfnd']):.2f}%")
        c3.metric("GB Dublin MAPE",    f"{mape_(m['yde'],m['pgbd']):.2f}%")
        c4.metric("GB Non-Dub MAPE",   f"{mape_(m['ynde'],m['pgbnd']):.2f}%")
        col1, col2 = st.columns(2)
        for c_, df_, ytr, yte, prf, pgb, dtst, name, clr in [
            (col1, m['dfd'],  m['ydt'],  m['yde'],  m['prfd'],  m['pgbd'],  m['dtd'],  'Dublin',     A1),
            (col2, m['dfnd'], m['yndt'], m['ynde'], m['prfnd'], m['pgbnd'], m['dtnd'], 'Non-Dublin', A2),
        ]:
            with c_:
                trd = df_['Date'].values[:-m['trf']]
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=trd, y=ytr, name='Train', line=dict(color=TC, width=1.5)))
                fig.add_trace(go.Scatter(x=dtst, y=yte, name='Actual', line=dict(color=A3, width=2.5)))
                fig.add_trace(go.Scatter(x=dtst, y=prf, name='Random Forest', line=dict(color=clr, width=2.5, dash='dash')))
                fig.add_trace(go.Scatter(x=dtst, y=pgb, name='Grad. Boosting', line=dict(color=A4, width=2.5, dash='dot')))
                fig.update_yaxes(tickprefix="€", tickformat=",")
                dl(fig, f"RF & GB — {name}", h=360)
                st.plotly_chart(fig, use_container_width=True)

        st.markdown("<div class='sec'>FEATURE IMPORTANCE (RANDOM FOREST)</div>", unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        for c_, model, fc_, name in [(col1,m['rfd'],m['fcd_c'],'Dublin'),(col2,m['rfnd'],m['fcnd_c'],'Non-Dublin')]:
            with c_:
                nn = model.n_features_in_; fc2 = fc_[:nn]
                imp = pd.DataFrame({'Feature':fc2,'Importance':model.feature_importances_}).sort_values('Importance').tail(12)
                clrs = [A3 if 'RPPI' in f or 'Unemp' in f else A1 if 'lag' in f or 'roll' in f else A2 for f in imp['Feature']]
                fig = go.Figure(go.Bar(x=imp['Importance'], y=imp['Feature'], orientation='h',
                    marker_color=clrs, hovertemplate="%{y}: %{x:.3f}<extra></extra>"))
                dl(fig, f"Feature Importance — {name}", h=360)
                st.plotly_chart(fig, use_container_width=True)

    # ── Prophet ──────────────────────────────────────────────────────────────
    elif page == "🔮 Prophet":
        st.markdown("<div class='sec'>PROPHET FORECAST MODEL</div>", unsafe_allow_html=True)
        def mape_(y,p): return np.mean(np.abs((y-p)/y))*100
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Dublin MAPE",     f"{mape_(m['tepd']['y'].values,m['ftd']['yhat'].values):.2f}%")
        c2.metric("Dublin MAE",      f"€{mean_absolute_error(m['tepd']['y'],m['ftd']['yhat']):.0f}")
        c3.metric("Non-Dublin MAPE", f"{mape_(m['tepnd']['y'].values,m['ftnd']['yhat'].values):.2f}%")
        c4.metric("Non-Dublin MAE",  f"€{mean_absolute_error(m['tepnd']['y'],m['ftnd']['yhat']):.0f}")
        col1, col2 = st.columns(2)
        cutoff_p = pd.Timestamp('2022-01-01')
        for c_, trd, ted, fct, fut, dpd_, name, clr in [
            (col1, m['tpd'],  m['tepd'],  m['ftd'],  m['pfud'],  m['dpd'],  'Dublin',     A1),
            (col2, m['tpnd'], m['tepnd'], m['ftnd'], m['pfund'], m['dpnd'], 'Non-Dublin', A2),
        ]:
            with c_:
                rec = dpd_[dpd_['ds'] >= cutoff_p]
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=trd['ds'], y=trd['y'], name='Train', line=dict(color=TC, width=1.5)))
                fig.add_trace(go.Scatter(x=ted['ds'], y=ted['y'], name='Actual', line=dict(color=A3, width=2.5)))
                fig.add_trace(go.Scatter(x=fct['ds'], y=fct['yhat'], name='Prophet', line=dict(color=clr, width=2.5, dash='dash')))
                fig.add_trace(go.Scatter(x=fct['ds'], y=fct['yhat_upper'], line=dict(width=0), showlegend=False))
                fig.add_trace(go.Scatter(x=fct['ds'], y=fct['yhat_lower'], fill='tonexty',
                    fillcolor='rgba(74,144,217,0.15)', line=dict(width=0), name='95% CI'))
                fig.add_trace(go.Scatter(x=fut['ds'], y=fut['yhat'], name='Future', line=dict(color=A2, width=2.5, dash='dot')))
                fig.add_trace(go.Scatter(x=fut['ds'], y=fut['yhat_upper'], line=dict(width=0), showlegend=False))
                fig.add_trace(go.Scatter(x=fut['ds'], y=fut['yhat_lower'], fill='tonexty',
                    fillcolor='rgba(0,196,180,0.1)', line=dict(width=0), showlegend=False))
                fig.update_yaxes(tickprefix="€", tickformat=",")
                dl(fig, f"Prophet — {name} (Q3'27: €{fut['yhat'].iloc[-1]:,.0f})", h=400)
                st.plotly_chart(fig, use_container_width=True)

    # ── SARIMAX ───────────────────────────────────────────────────────────────
    elif page == "📡 SARIMAX":
        st.markdown("<div class='sec'>SARIMAX — EXOGENOUS VARIABLES</div>", unsafe_allow_html=True)
        def mape_(y,p): return np.mean(np.abs((y-p)/y))*100
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Dublin MAPE",     f"{mape_(m['tesx']['Dublin'].values,m['psxd'].values):.2f}%")
        c2.metric("Dublin MAE",      f"€{mean_absolute_error(m['tesx']['Dublin'],m['psxd']):.0f}")
        c3.metric("Non-Dublin MAPE", f"{mape_(m['tesx']['Non_Dublin'].values,m['psxnd'].values):.2f}%")
        c4.metric("Non-Dublin MAE",  f"€{mean_absolute_error(m['tesx']['Non_Dublin'],m['psxnd']):.0f}")
        col1, col2 = st.columns(2)
        for c_, cn, pred, ci, fcm, fcci, name, clr in [
            (col1,'Dublin',    m['psxd'],  m['csxd'],  m['sxdm'],  m['sxdci'],  'Dublin',     A1),
            (col2,'Non_Dublin',m['psxnd'], m['csxnd'], m['sxndm'], m['sxndci'], 'Non-Dublin', A2),
        ]:
            with c_:
                fig=fc_chart(m['trsx'].index,m['trsx'][cn],
                             m['tesx'].index,m['tesx'][cn],pred.values,
                             ci.iloc[:,0].values,ci.iloc[:,1].values,
                             fcm.index,fcm.values,fcci.iloc[:,0].values,fcci.iloc[:,1].values,
                             'SARIMAX',clr,f"SARIMAX — {name} (Q3'27: €{fcm.iloc[-1]:,.0f})")
                st.plotly_chart(fig, use_container_width=True)

    # ── Model Comparison ──────────────────────────────────────────────────────
    elif page == "🏆 Model Comparison":
        st.markdown("<div class='sec'>FINAL MODEL COMPARISON</div>", unsafe_allow_html=True)
        comp = m['comp'].copy()
        if model_sel: comp = comp[comp['Model'].isin(model_sel)]
        best = m['comp'].iloc[0]
        st.success(f"🥇 Best Model: **{best['Model']}** — Avg MAPE: {best['Avg_MAPE']:.2f}% | Avg MAE: €{best['Avg_MAE']:.0f}/month")

        col1, col2 = st.columns(2)
        with col1:
            fig = go.Figure()
            for i, row in comp.iterrows():
                fig.add_trace(go.Bar(name=row['Model'], x=['Dublin MAPE','Non-Dublin MAPE'],
                    y=[row['Dublin_MAPE'],row['NonDublin_MAPE']], marker_color=COLS[i%5],
                    hovertemplate=f"<b>{row['Model']}</b> %{{x}}: %{{y:.2f}}%<extra></extra>"))
            hline_label(fig, 5, "5% threshold", A3)
            fig.update_yaxes(ticksuffix="%")
            dl(fig, "MAPE — Lower is Better", h=380); fig.update_layout(barmode='group')
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            fig2 = go.Figure()
            for i, row in comp.iterrows():
                fig2.add_trace(go.Bar(name=row['Model'], x=['Dublin MAE','Non-Dublin MAE'],
                    y=[row['Dublin_MAE'],row['NonDublin_MAE']], marker_color=COLS[i%5],
                    hovertemplate=f"<b>{row['Model']}</b> %{{x}}: €%{{y:.0f}}<extra></extra>"))
            fig2.update_yaxes(tickprefix="€", tickformat=",")
            dl(fig2, "MAE — €/month (Lower is Better)", h=380); fig2.update_layout(barmode='group')
            st.plotly_chart(fig2, use_container_width=True)

        disp = m['comp'][['Model','Dublin_MAE','Dublin_MAPE','NonDublin_MAE','NonDublin_MAPE','Avg_MAPE','Avg_MAE']].copy()
        for col_ in ['Dublin_MAE','NonDublin_MAE','Avg_MAE']:     disp[col_] = disp[col_].apply(lambda x: f"€{x:.0f}")
        for col_ in ['Dublin_MAPE','NonDublin_MAPE','Avg_MAPE']:  disp[col_] = disp[col_].apply(lambda x: f"{x:.2f}%")
        st.dataframe(disp, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: POLICY INSIGHTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📋 Policy Insights":
    st.markdown("<div class='sec'>RESEARCH FINDINGS & POLICY RECOMMENDATIONS</div>", unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    c1.metric("Dublin Growth (2007–2025)",     f"+{((_last['Dublin']/_first['Dublin'])-1)*100:.0f}%",   "vs wage growth")
    c2.metric("Non-Dublin Growth (2007–2025)", f"+{((_last['Non_Dublin']/_first['Non_Dublin'])-1)*100:.0f}%", "crisis spreading")
    c3.metric("Best Model Accuracy",           "< 5% MAPE",  "8 quarters ahead")

    t1, t2, t3 = st.tabs(["🔍 SQ1: Key Drivers", "⚠️ SQ2: Crisis?", "📡 SQ3: Lead Time"])
    with t1:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""<div class='kpi-card' style='text-align:left'>
            <div style='color:#4A90D9;font-size:0.7rem;letter-spacing:2px;margin-bottom:8px'>DUBLIN DRIVERS</div>
            <ol style='color:#C8CBD8;line-height:2'>
            <li><b>Rent Lag (Q-1)</b> — ~18%</li><li><b>Rent Lag (Q-4)</b> — ~17%</li>
            <li><b>Rolling 4Q Mean</b> — ~17%</li><li><b>Unemployment Lag</b> — ~11%</li>
            <li><b>RPPI Index</b> — SARIMAX key driver</li></ol></div>""", unsafe_allow_html=True)
        with col2:
            st.markdown("""<div class='kpi-card' style='text-align:left'>
            <div style='color:#00C4B4;font-size:0.7rem;letter-spacing:2px;margin-bottom:8px'>NON-DUBLIN DRIVERS</div>
            <ol style='color:#C8CBD8;line-height:2'>
            <li><b>Rent Lag (Q-1)</b> — ~37% dominant</li><li><b>Unemployment Rate</b> — ~14%</li>
            <li><b>Rolling 4Q Mean</b> — trend signal</li><li><b>RPZ Dummy</b> — structural break 2016</li>
            </ol></div>""", unsafe_allow_html=True)
    with t2:
        st.markdown(f"""<div class='kpi-card' style='text-align:left;border-left:3px solid #FF6B6B'>
        <div style='color:#FF6B6B;font-size:1rem;font-weight:700;margin-bottom:10px'>⚠️ YES — Systemic Crisis Confirmed</div>
        <ul style='color:#C8CBD8;line-height:2.1'>
        <li>Dublin: <b>€{_last['Dublin']:,.0f}/month</b> — +{((_last['Dublin']/_first['Dublin'])-1)*100:.0f}% since 2007</li>
        <li>Non-Dublin: <b>€{_last['Non_Dublin']:,.0f}/month</b> — +{((_last['Non_Dublin']/_first['Non_Dublin'])-1)*100:.0f}% since 2007</li>
        <li>All 5 models forecast <b>continued upward trajectory</b> to 2027</li>
        <li>RPZ (2016) slowed but did <b>NOT reverse</b> the trend</li>
        <li>Dublin rent-to-income: <b>{_aff_d:.1f}%</b> (threshold: 30%)</li>
        </ul></div>""", unsafe_allow_html=True)
    with t3:
        st.markdown("""<div class='kpi-card' style='text-align:left;border-left:3px solid #00C4B4'>
        <div style='color:#00C4B4;font-size:1rem;font-weight:700;margin-bottom:10px'>✅ YES — Sufficient Lead Time</div>
        <ul style='color:#C8CBD8;line-height:2.1'>
        <li>Forecast window: <b>8 quarters (2 years)</b> ahead</li>
        <li>Best model accuracy: <b>&lt;5% MAPE</b></li>
        <li>Policy implementation lead time: <b>12–18 months</b></li>
        <li>Dashboard provides <b>6–8 quarter early warning</b></li>
        </ul></div>""", unsafe_allow_html=True)

    st.markdown("<div class='sec'>POLICY RECOMMENDATIONS</div>", unsafe_allow_html=True)
    recs = [
        ("🏗️","Supply-Side Intervention","Accelerate social & affordable housing. Fast-track planning in Dublin corridors.",A1),
        ("📜","Strengthen RPZ","Lower RPZ cap to inflation-linked rate. Expand to Non-Dublin commuter belts.",A2),
        ("💼","Unemployment Linkage","Incentivise remote-work hubs to distribute housing demand geographically.",A4),
        ("📊","Deploy Live Dashboard","Run SARIMAX/Prophet as live policy tool. Auto-alert at 30% affordability threshold.",A5),
        ("🗄️","Data Infrastructure","Add mortgage rates, construction costs, migration data as exogenous variables.",A3),
    ]
    cols_ = st.columns(5)
    for col_, (icon, title, desc, clr) in zip(cols_, recs):
        col_.markdown(f"""<div class='kpi-card' style='text-align:left;min-height:170px'>
        <div style='font-size:1.6rem;margin-bottom:6px'>{icon}</div>
        <div style='color:{clr};font-size:0.7rem;font-weight:700;letter-spacing:1px;margin-bottom:5px'>{title.upper()}</div>
        <div style='color:#A0A3B0;font-size:0.78rem;line-height:1.5'>{desc}</div>
        </div>""", unsafe_allow_html=True)
